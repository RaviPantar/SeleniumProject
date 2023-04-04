import openpyxl

workbook = openpyxl.load_workbook("C:\\Users\\UNIFY\\Documents\\Book1.xlsx")
sheet = workbook.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=2, column=2).value = "Rahul"

print(sheet.cell(row=2, column=2).value)

print(sheet['A3'].value)

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)

